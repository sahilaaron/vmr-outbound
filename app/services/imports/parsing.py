"""Shared spreadsheet parsing for the two authorized import formats (CSV, XLSX).

Both formats are parsed into one neutral shape — sheets of header + verbatim
string rows — so every downstream stage (mapping, validation, normalization,
deduplication, provenance, suppression, persistence) runs on exactly one code
path. Business rules are never duplicated per format.

Boundaries (product rule): ``.csv`` and ``.xlsx`` only. ``.xls``, Google Sheets
direct import, and other formats are rejected visibly at this layer.

XLSX specifics:

* Parsed with ``openpyxl`` in read-only mode; cell values are rendered to
  strings verbatim-as-displayed (numbers keep their repr, dates use ISO format).
* Sheet name, sheet index, and the original per-sheet row number are preserved
  on every row.
* A malformed workbook (not a real ``.xlsx``) or an empty workbook (no sheet
  with a header) is a visible, batch-level failure — never a silent success.

A CSV file is represented as a single sheet with index 0 and no sheet name, so
flat files flow through the same pipeline unchanged.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook

CSV_PARSER_VERSION = "csv-2"
XLSX_PARSER_VERSION = "xlsx-1"

SUPPORTED_EXTENSIONS = (".csv", ".xlsx")

_UNMAPPED_KEY = "_unmapped"

#: Largest single CSV cell this reader will accept, in characters.
#:
#: Python's default is 128 KB, which a real Apollo export can exceed — the
#: Keywords and Technologies columns routinely run to kilobytes and occasionally
#: much further. Left at the default, an ordinary export raised ``_csv.Error``,
#: which is not a ``MalformedFileError`` and so escaped every handler on the way
#: out as a bare 500. Raised here to a bound of its own rather than removed:
#: unbounded would mean one hostile cell could hold the whole upload ceiling in
#: a single Python string, and 4 MiB is far above any plausible cell while
#: staying well inside the 25 MB file limit.
MAX_CSV_FIELD_CHARS = 4 * 1024 * 1024

csv.field_size_limit(MAX_CSV_FIELD_CHARS)

#: Most physical columns a sheet may have, counted by POSITION.
#:
#: Positions, not names. The operator-facing ceiling was applied to the filtered
#: list of non-blank headers, so a file with the four required columns and five
#: hundred blank ones passed a five-hundred-and-twelve-column limit at five
#: hundred and sixteen columns — and every one of those blanks becomes a Python
#: object on every row before the limit is consulted. Enforced here, at the
#: moment the header row is read, so the expansion is never materialized.
MAX_PHYSICAL_COLUMNS = 512

_CSV_FIELD_LIMIT_MESSAGE = (
    "The file contains a single cell larger than this import accepts "
    f"({MAX_CSV_FIELD_CHARS // (1024 * 1024)} MB). Split the value or remove the "
    "column and re-export the file."
)

_CSV_QUOTING_MESSAGE = (
    "The file has a quoting error: a quoted value is never closed, or a quotation "
    "mark appears in the middle of an unquoted value. Open the file in a "
    "spreadsheet application and re-export it as CSV so the quoting is written "
    "correctly."
)


def _csv_error_message(exc: Exception) -> str:
    """Say which of the two CSV failures happened, accurately.

    One message used to cover both and asserted that an unclosed quote was
    rejected — which was not true before the reader was made strict, and is a
    claim worth keeping honest now that it is.
    """

    return _CSV_FIELD_LIMIT_MESSAGE if "field limit" in str(exc) else _CSV_QUOTING_MESSAGE


def positional_key(index: int) -> str:
    """A stable durable name for a physical column that has no header.

    One-based, matching the column letters an operator sees in a spreadsheet
    application after converting them. Needed because ``import_rows.raw_data``
    is the immutable record of the file and a JSON object needs a key: before
    this, every headerless column's value was joined into a single ``_unmapped``
    string, so two blank columns holding ``left`` and ``right`` became one cell
    and nobody could say afterwards which position supplied which value.

    Deliberately not a human-looking header. Inventing ``Column E`` would put a
    name in the record that the file does not contain.
    """

    return f"(column {index + 1})"


def _check_width(columns: Sequence[str], *, sheet_name: str | None) -> None:
    """Refuse a sheet that is wider than the reader will process."""

    if len(columns) <= MAX_PHYSICAL_COLUMNS:
        return
    where = f"Sheet {sheet_name!r}" if sheet_name else "The file"
    raise MalformedFileError(
        f"{where} has {len(columns):,} columns, which is more than the "
        f"{MAX_PHYSICAL_COLUMNS:,}-column limit. Blank and unnamed columns count: "
        "delete the unused columns to the right of your data and re-export."
    )


class UnsupportedFormatError(Exception):
    """Raised for a file that is not one of the authorized formats."""


class MalformedFileError(Exception):
    """Raised when a file cannot be parsed as its claimed format."""


@dataclass(frozen=True)
class ParsedRow:
    """One verbatim data row, by header name and by column position.

    Both, because neither alone is enough. ``raw`` is what gets stored durably
    in ``import_rows.raw_data`` as JSONB, and a JSON object cannot hold two
    entries under one key — so a file with two columns literally named ``Email``
    has to lose one of them there. ``cells`` is positional and loses nothing,
    and it is what the schema reading uses.

    The rule where they disagree is that **the first column of a repeated name
    wins** in ``raw``. That is the contract :func:`app.services.imports.apollo
    .detect_schema` states, and before this pair existed the dict silently did
    the opposite: the later column overwrote the earlier one, so the reading
    took a value the preview had just told the operator was not applied.
    """

    sheet_index: int
    sheet_name: str | None
    row_number: int  # original per-sheet data-row number (header excluded)
    raw: dict[str, str]
    #: Values in file column order, aligned index-for-index with
    #: :attr:`SheetInfo.columns`. Shorter than the header when the row is short.
    cells: tuple[str, ...] = ()


@dataclass(frozen=True)
class SheetInfo:
    """Inspection summary of one sheet (for the workbook-inspection step)."""

    index: int
    name: str | None
    header: list[str]
    data_row_count: int
    #: The header row verbatim, including empty and repeated names, so that
    #: position ``i`` here names position ``i`` in every row's ``cells``.
    #: :attr:`header` is the display list and drops the empties, which is why it
    #: cannot be used for alignment.
    columns: tuple[str, ...] = ()


@dataclass
class ParsedFile:
    """A parsed CSV or XLSX file in the neutral shared shape."""

    source_format: str  # "csv" | "xlsx"
    parser_version: str
    sheets: list[SheetInfo] = field(default_factory=list)
    rows: list[ParsedRow] = field(default_factory=list)

    def sheet(self, index: int) -> SheetInfo | None:
        for info in self.sheets:
            if info.index == index:
                return info
        return None

    def rows_for_sheets(self, indexes: list[int] | None) -> list[ParsedRow]:
        """Rows restricted to the selected sheets (all rows when None)."""

        if indexes is None:
            return list(self.rows)
        wanted = set(indexes)
        return [row for row in self.rows if row.sheet_index in wanted]


def detect_format(filename: str | None) -> str:
    """Return "csv" or "xlsx" from the filename, or raise for anything else."""

    name = (filename or "").strip().lower()
    if name.endswith(".csv"):
        return "csv"
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".xls"):
        raise UnsupportedFormatError(
            "Legacy .xls workbooks are not supported. Save the file as .xlsx "
            "(or export a .csv) and upload again."
        )
    raise UnsupportedFormatError(
        "Unsupported file type. The import accepts .csv and .xlsx files only."
    )


def _cell_to_text(value: Any) -> str:
    """Render one cell value to its verbatim string form."""

    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        # A date-only cell arrives as midnight; render it as the operator saw it.
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def parse_csv(content: bytes) -> ParsedFile:
    """Parse CSV bytes into the shared shape (single sheet, index 0)."""

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise MalformedFileError(
            "The file is not readable as UTF-8 text. Re-export the CSV with "
            "UTF-8 encoding and upload again."
        ) from exc

    # ``csv.reader`` rather than ``DictReader``: the dict is built here so that a
    # repeated header name keeps its FIRST column's value instead of being
    # overwritten by the last, and so the positional reading survives at all.
    # ``strict=True`` so a malformed quote is an error rather than a silent
    # transformation. Without it, ``"A"da`` was quietly read as ``Ada`` — the
    # immutable raw row then recorded text the file did not contain — and an
    # unterminated quote swallowed the rest of the line into one cell.
    reader = csv.reader(io.StringIO(text), strict=True)
    try:
        first = next(reader, None)
    except csv.Error as exc:
        raise MalformedFileError(_csv_error_message(exc)) from exc
    columns: tuple[str, ...] = tuple(first or ())
    # Before a single data row is materialized.
    _check_width(columns, sheet_name=None)

    rows: list[ParsedRow] = []
    row_number = 0
    try:
        for values in reader:
            cells = tuple(str(v) if v is not None else "" for v in values)
            if not any(v.strip() for v in cells):
                continue  # skip fully-empty lines
            cleaned: dict[str, str] = {}
            for position, value in enumerate(cells):
                if position < len(columns) and columns[position]:
                    cleaned.setdefault(columns[position], value)
                elif value.strip():
                    # A headerless or overflow column keeps its own position, so
                    # the durable record can still say which cell held what.
                    cleaned[positional_key(position)] = value
            # Every declared column is present even when the row is short, so a
            # ragged file reads as empty cells rather than as missing fields.
            for name in columns:
                if name:
                    cleaned.setdefault(name, "")
            row_number += 1
            rows.append(
                ParsedRow(
                    sheet_index=0,
                    sheet_name=None,
                    row_number=row_number,
                    raw=cleaned,
                    cells=cells,
                )
            )
    except csv.Error as exc:
        raise MalformedFileError(_csv_error_message(exc)) from exc

    header = [h for h in columns if h and h != _UNMAPPED_KEY]
    parsed = ParsedFile(source_format="csv", parser_version=CSV_PARSER_VERSION)
    parsed.sheets = [
        SheetInfo(index=0, name=None, header=header, data_row_count=len(rows), columns=columns)
    ]
    parsed.rows = rows
    return parsed


def parse_xlsx(content: bytes) -> ParsedFile:
    """Parse XLSX bytes into the shared shape, one entry per worksheet.

    The first non-empty row of each sheet is its header. Sheets with no header
    row contribute no rows but still appear in the inspection summary (with an
    empty header) so the operator can see they were found and skipped.
    """

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, KeyError, OSError, ValueError) as exc:
        raise MalformedFileError(
            "The file could not be opened as an .xlsx workbook. It may be "
            "corrupted, password-protected, or mislabelled. Re-export it from "
            "the source application and upload again."
        ) from exc

    try:
        parsed = ParsedFile(source_format="xlsx", parser_version=XLSX_PARSER_VERSION)
        if not workbook.sheetnames:
            raise MalformedFileError("The workbook contains no sheets.")

        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet_name]
            header: list[str] = []
            data_row_count = 0
            row_number = 0
            for values in worksheet.iter_rows(values_only=True):
                texts = [_cell_to_text(v) for v in values]
                if not any(t.strip() for t in texts):
                    continue  # skip fully-empty rows (including leading ones)
                if not header:
                    header = [t.strip() for t in texts]
                    _check_width(header, sheet_name=sheet_name)
                    continue
                row_number += 1
                cells = tuple(texts)
                raw: dict[str, str] = {}
                for position, text in enumerate(texts):
                    if position < len(header) and header[position]:
                        # First column of a repeated name wins, matching the CSV
                        # reader and the documented detection contract.
                        raw.setdefault(header[position], text)
                    elif text.strip():
                        raw[positional_key(position)] = text
                for name in header:
                    if name:
                        raw.setdefault(name, "")
                data_row_count += 1
                parsed.rows.append(
                    ParsedRow(
                        sheet_index=sheet_index,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        raw=raw,
                        cells=cells,
                    )
                )
            parsed.sheets.append(
                SheetInfo(
                    index=sheet_index,
                    name=sheet_name,
                    header=[h for h in header if h],
                    data_row_count=data_row_count,
                    columns=tuple(header),
                )
            )

        if all(not sheet.header for sheet in parsed.sheets):
            raise MalformedFileError(
                "The workbook is empty: no sheet contains a header row. Add a "
                "header row naming the contact columns and upload again."
            )
        return parsed
    finally:
        workbook.close()


def parse_file(content: bytes, filename: str | None) -> ParsedFile:
    """Parse an uploaded file by its extension into the shared shape."""

    file_format = detect_format(filename)
    if not content:
        raise MalformedFileError("The uploaded file is empty (0 bytes).")
    if file_format == "csv":
        return parse_csv(content)
    return parse_xlsx(content)
