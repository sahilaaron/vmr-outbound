"""Builders for Apollo-shaped contact exports (IMP-001 tests).

Files are built here rather than committed as binary fixtures for two reasons.
A committed ``.xlsx`` is opaque in review — nobody can see from a diff what a
test is actually asserting about — and a workbook that has to exercise formulas,
hidden sheets or a second worksheet is easier to state in code than to describe
in a filename.
"""

from __future__ import annotations

import csv
import io
import uuid
from typing import Any

from app.models.campaign import Campaign
from app.models.enums import CampaignStatus
from openpyxl import Workbook
from sqlalchemy.orm import Session

#: The Apollo contact export header, in the order the vendor emits it. Tests that
#: care about order-independence shuffle or subset this deliberately.
APOLLO_HEADER: tuple[str, ...] = (
    "First Name",
    "Last Name",
    "Title",
    "Seniority",
    "Departments",
    "Sub Departments",
    "Person Linkedin Url",
    "City",
    "State",
    "Country",
    "Phone",
    "Mobile Phone",
    "Corporate Phone",
    "Email",
    "Result",
    "Email Status",
    "Primary Email Source",
    "Primary Email Verification Source",
    "Primary Email Catch-all Status",
    "Primary Email Last Verified At",
    "Secondary Email",
    "Secondary Email Source",
    "Secondary Email Status",
    "Secondary Email Verification Source",
    "Secondary Email Last Verified At",
    "Tertiary Email",
    "Tertiary Email Source",
    "Tertiary Email Status",
    "Tertiary Email Verification Source",
    "Tertiary Email Last Verified At",
    "Company Name",
    "Company Name for Emails",
    "Website",
    "Company Linkedin Url",
    "Company Address",
    "Company City",
    "Company State",
    "Company Country",
    "# Employees",
    "Industry",
    "Keywords",
    "Technologies",
    "Annual Revenue",
    "Apollo Contact Id",
    "Apollo Account Id",
    "Apollo Record Id",
)

#: The minimum a file needs for the schema to be recognized at all.
MINIMAL_HEADER: tuple[str, ...] = ("First Name", "Last Name", "Company Name", "Email")


def row(**overrides: Any) -> dict[str, str]:
    """One plausible Apollo row, overridable by header name.

    Defaults describe a straightforward person at a company whose website and
    email domain agree, so a test that cares about one specific hazard only has
    to state that hazard.

    The per-person identifiers — the Apollo ids and the LinkedIn profile URL —
    are DERIVED from the address unless a test overrides them. Fixed defaults
    would have made every row in a multi-row file the same person by Apollo id,
    which is a real outcome the importer must produce and a terrible accident to
    produce by default. A test that wants that collision now has to ask for it.
    """

    base: dict[str, str] = dict.fromkeys(APOLLO_HEADER, "")
    base.update(
        {
            "First Name": "Ada",
            "Last Name": "Lovelace",
            "Title": "Head of Analytical Engines",
            "Seniority": "head",
            "Departments": "engineering",
            "Person Linkedin Url": "https://www.linkedin.com/in/ada-lovelace",
            "City": "London",
            "State": "England",
            "Country": "United Kingdom",
            "Email": "ada@engines.example",
            "Result": "Verified",
            "Email Status": "Valid",
            "Primary Email Source": "Apollo",
            "Primary Email Verification Source": "Apollo Verification",
            "Primary Email Catch-all Status": "Not Catch-all",
            "Primary Email Last Verified At": "2026-05-01T09:30:00Z",
            "Company Name": "Analytical Engines",
            "Company Name for Emails": "Analytical Engines",
            "Website": "https://engines.example",
            "Company Linkedin Url": "https://www.linkedin.com/company/analytical-engines",
            "Company City": "London",
            "Company Country": "United Kingdom",
            "# Employees": "120",
            "Industry": "software",
            "Apollo Account Id": "apollo-account-1",
        }
    )
    for key, value in overrides.items():
        base[key] = value

    handle = (base.get("Email") or "unknown").split("@", 1)[0].replace(".", "-") or "unknown"
    base.setdefault("Apollo Contact Id", "")
    if not overrides.get("Apollo Contact Id") and not base["Apollo Contact Id"]:
        base["Apollo Contact Id"] = f"apollo-contact-{handle}"
    if not overrides.get("Apollo Record Id") and not base["Apollo Record Id"]:
        base["Apollo Record Id"] = f"apollo-record-{handle}"
    if "Person Linkedin Url" not in overrides:
        base["Person Linkedin Url"] = f"https://www.linkedin.com/in/{handle}"
    return base


def csv_bytes(
    rows: list[dict[str, str]],
    *,
    header: tuple[str, ...] = APOLLO_HEADER,
    encoding: str = "utf-8",
    bom: bool = False,
    delimiter: str = ",",
) -> bytes:
    """Render rows as CSV exactly as a spreadsheet application would."""

    buffer = io.StringIO()
    writer = csv.DictWriter(
        buffer, fieldnames=list(header), extrasaction="ignore", delimiter=delimiter
    )
    writer.writeheader()
    for entry in rows:
        writer.writerow({key: entry.get(key, "") for key in header})
    text = buffer.getvalue()
    encoded = text.encode(encoding)
    return (b"\xef\xbb\xbf" + encoded) if bom else encoded


def xlsx_bytes(
    sheets: dict[str, tuple[tuple[str, ...], list[dict[str, str]]]],
) -> bytes:
    """Render one workbook, one entry per worksheet."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, (header, rows) in sheets.items():
        worksheet = workbook.create_sheet(title=title)
        worksheet.append(list(header))
        for entry in rows:
            worksheet.append([entry.get(column, "") for column in header])
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def xlsx_positional_bytes(sheet_name: str, header: tuple[str, ...], rows: list[list[str]]) -> bytes:
    """A workbook whose rows are given BY POSITION rather than by header name.

    :func:`xlsx_bytes` takes header-keyed dictionaries, which cannot express two
    different values under two blank headers — the dict collapses them before
    openpyxl ever sees them. Any test about repeated blank headers written
    against that helper is testing the helper, not the parser.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(list(header))
    for row in rows:
        worksheet.append(list(row))
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def xlsx_with_formula(header: tuple[str, ...], rows: list[dict[str, str]]) -> bytes:
    """A workbook whose first data row carries a real, uncached formula cell.

    ``openpyxl`` is opened with ``data_only=True``, which reads the value Excel
    last cached for a formula. A workbook written by openpyxl has never been
    opened by Excel, so there is no cached value and the cell reads as empty —
    which is the correct, safe behaviour and is what the test asserts. What must
    never happen is the formula being evaluated.
    """

    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = workbook.create_sheet(title="Contacts")
    worksheet.append(list(header))
    for index, entry in enumerate(rows):
        worksheet.append([entry.get(column, "") for column in header])
        if index == 0:
            worksheet.cell(row=2, column=len(header) + 1, value="=1+1")
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def make_campaign(
    session: Session, *, name: str | None = None, execution: bool = False
) -> Campaign:
    """A campaign to import into. Execution is off unless a test asks for it."""

    campaign = Campaign(
        name=name or f"Import campaign {uuid.uuid4()}",
        status=CampaignStatus.ACTIVE,
        execution_enabled=execution,
    )
    session.add(campaign)
    session.flush()
    return campaign
